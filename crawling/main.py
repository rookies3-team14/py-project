import os
import pandas as pd
from openpyxl import load_workbook
from get_recruitment_text import get_saramin_recruitment_text

path = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(path, "../data/excel/RecruitmentNotice.xlsx")
new_file_path = os.path.join(path, "../data/excel/stack_candidate.xlsx")

# 엑셀 파일 없거나 비어있으면 초기화
if not os.path.exists(new_file_path) or os.path.getsize(new_file_path) == 0:
    dummy_df = pd.DataFrame(
        columns=["companyName", "title", "recruitUrl", "annual", "text"])
    dummy_df.to_excel(new_file_path, index=False)
    print("📄 새 파일 생성 완료 (헤더만 있음)")

df = pd.read_excel(file_path)
df_filtered = df.iloc[339:].reset_index(drop=True)
json_list = df_filtered.to_dict(orient="records")

for item in json_list:
    updated = get_saramin_recruitment_text(item, item["recruitUrl"])
    if not updated:
        continue

    new_row = pd.DataFrame([updated])

    # 기존 엑셀 파일에서 시작 행 계산
    book = load_workbook(new_file_path)
    sheet = book.active
    start_row = sheet.max_row

    with pd.ExcelWriter(new_file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        new_row.to_excel(writer, index=False, header=False, startrow=start_row)
        print(f"✅ 저장 완료: {updated['companyName']}")
